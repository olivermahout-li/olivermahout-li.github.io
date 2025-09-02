import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # Graceful message if dependency missing
    print("Please install openpyxl: pip install openpyxl", file=sys.stderr)
    raise


WORKSPACE_ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX_PATH = WORKSPACE_ROOT / "library_update_this_file.xlsx"
DEFAULT_HTML_PATH = WORKSPACE_ROOT / "index.html"


def read_rows_as_dicts(ws) -> List[Dict[str, str]]:
    """Read a worksheet into a list of dicts using the first non-empty row as headers.

    Empty rows are skipped. Cell values are converted to strings (strip spaces)."""
    rows = list(ws.iter_rows(values_only=True))
    # Find header row (first non-empty row)
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
            header_idx = i
            break
    if header_idx is None:
        return []

    headers_raw = ["" if c is None else str(c).strip() for c in rows[header_idx]]
    headers = []
    for idx, h in enumerate(headers_raw):
        if h == "":
            headers.append(f"col_{idx}")
        else:
            headers.append(h)

    data_dicts: List[Dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        if row is None:
            continue
        # Skip entirely empty rows
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        entry: Dict[str, str] = {}
        for key, cell in zip(headers, row):
            entry[key] = "" if cell is None else str(cell).strip()
        data_dicts.append(entry)
    return data_dicts


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def find_col(entry: Dict[str, str], possible: List[str]) -> Optional[str]:
    """Find a key in entry that matches any name in possible (case-insensitive)."""
    lower_map = {k.lower(): k for k in entry.keys()}
    for p in possible:
        if p.lower() in lower_map:
            return lower_map[p.lower()]
    # try partial contains
    for k in entry.keys():
        kl = k.lower()
        if any(p.lower() in kl for p in possible):
            return k
    return None


NAME_HIGHLIGHT_PATTERN = re.compile(r"\b(Yuchen\s+Li|Y\.?\s*Li)\b", re.IGNORECASE)


def highlight_name_globally(text: str) -> str:
    return NAME_HIGHLIGHT_PATTERN.sub(r"<font class=\"deepblue\">\1</font>", text)


def highlight_authors(authors: str, author_order_value: Optional[str]) -> str:
    """Highlight the author's name based on author_order if provided; otherwise global match.

    Strategy:
    - Split authors by comma to preserve the common style in index.html
    - If author_order is valid (1-based), wrap that author token entirely with deepblue font
    - Otherwise, highlight by name pattern globally
    """
    if not authors:
        return authors
    # Try parse author_order
    author_index: Optional[int] = None
    if author_order_value:
        try:
            author_index = int(str(author_order_value).strip())
            if author_index < 1:
                author_index = None
        except Exception:
            author_index = None

    # If we have a valid index, try to highlight that token
    if author_index is not None:
        tokens = [t.strip() for t in authors.split(',')]
        if 1 <= author_index <= len(tokens):
            tokens = [
                (f"<font class=\"deepblue\">{tok}</font>" if i + 1 == author_index else tok)
                for i, tok in enumerate(tokens)
            ]
            return ", ".join(tokens)

    # Fallback: highlight by name pattern
    return highlight_name_globally(authors)


def build_news_items(rows: List[Dict[str, str]]) -> str:
    items: List[str] = []
    for r in rows:
        date_key = find_col(r, ["date", "日期", "时间"])
        text_key = find_col(r, ["text", "content", "描述", "description", "desc", "内容", "news"]) or next(
            iter(r.keys()),
            None,
        )
        date_val = (r.get(date_key, "") if date_key else "").strip()
        text_val = (r.get(text_key, "") if text_key else "").strip()
        if not text_val:
            # Fallback: join all non-empty values
            text_val = ", ".join(v for v in r.values() if v)
        if date_val:
            items.append(f"<li><i>{date_val}</i> : {text_val}</li>")
        else:
            items.append(f"<li>{text_val}</li>")
        items.append("<br>")
    return "\n".join(items)


def compose_pub_main_text(r: Dict[str, str]) -> str:
    """Compose the bibliographic line with required order:
    authors, "title", italic(container), pages, time, doi
    """
    raw_authors = r.get(find_col(r, ["authors", "author", "作者"]) or "", "").strip()
    author_order_key = find_col(r, ["author_order", "作者序", "作者次序", "次序"]) or ""
    author_order_val = r.get(author_order_key, "").strip() if author_order_key else ""
    authors = highlight_authors(raw_authors, author_order_val)

    title = r.get(find_col(r, ["title", "题目"]) or "", "").strip()
    container = r.get(
        find_col(r, ["journal", "conference", "book", "venue", "容器", "刊物", "期刊", "会议"]) or "",
        "",
    ).strip()

    pages = r.get(find_col(r, ["pages", "pp", "页"]) or "", "").strip()

    # Prefer explicit time column if present, else combine month + year
    time_key = find_col(r, ["time", "日期", "时间"]) or ""
    time_val = r.get(time_key, "").strip() if time_key else ""
    if not time_val:
        month = r.get(find_col(r, ["month"]) or "", "").strip()
        year = r.get(find_col(r, ["year", "年份"]) or "", "").strip()
        time_val = ", ".join([v for v in [month, year] if v])

    doi_key = find_col(r, ["doi", "doa"]) or ""
    doi_val = r.get(doi_key, "").strip() if doi_key else ""

    parts: List[str] = []
    if authors:
        parts.append(authors)
    if title:
        parts.append(f'"{title}"')
    if container:
        parts.append(f"<i>{container}</i>")
    if pages:
        parts.append(pages)
    if time_val:
        parts.append(time_val)
    if doi_val:
        parts.append(f"doi: {doi_val}")

    main = ", ".join(parts).strip()
    if main and not main.endswith("."):
        main += "."
    return main


def build_pub_item(r: Dict[str, str]) -> str:
    main_text = compose_pub_main_text(r)
    note_key = find_col(r, ["note", "备注"])
    note_val = (r.get(note_key, "") if note_key else "").strip()
    
    note_link_key = find_col(r, ["note_link", "备注链接", "note_url"])
    note_link_val = (r.get(note_link_key, "") if note_link_key else "").strip()

    link_key = find_col(r, ["url", "link", "链接", "地址"]) or ""
    link_val = r.get(link_key, "").strip() if link_key else ""

    doi_key = find_col(r, ["doi", "doa"]) or ""
    doi_val = r.get(doi_key, "").strip() if doi_key else ""

    extras: List[str] = []
    if note_val:
        if note_link_val:
            # Make note a hyperlink if note_link exists
            extras.append(f"<font class=\"red\">(<a href='{note_link_val}' class=\"red\">{note_val}</a>)</font>")
        else:
            # Regular note without link
            extras.append(f"<font class=\"red\">({note_val})</font>")

    link_html = ""
    if link_val:
        link_html = f"<br><a href='{link_val}' class=\"deepgrey\">{link_val}</a>"
    elif doi_val:
        doi_url = f"https://doi.org/{doi_val}"
        link_html = f"<br><a href='{doi_url}' class=\"deepgrey\">{doi_url}</a>"

    extras_text = (" " + " ".join(extras)) if extras else ""
    return f"<li>{main_text}{extras_text}{link_html}</li>"


def build_pub_item_with_trailing_br(r: Dict[str, str]) -> str:
    return build_pub_item(r) + "\n<br>"


def build_book_chapter_item(r: Dict[str, str]) -> str:
    """Build book chapter with specific format and optional image."""
    # Extract specific fields for book chapters
    raw_authors = r.get(find_col(r, ["authors", "author", "作者"]) or "", "").strip()
    author_order_key = find_col(r, ["author_order", "作者序", "作者次序", "次序"]) or ""
    author_order_val = r.get(author_order_key, "").strip() if author_order_key else ""
    authors = highlight_authors(raw_authors, author_order_val)
    
    chapter_title = r.get(find_col(r, ["chapter_title", "chapter", "章节题目"]) or "", "").strip()
    book_title = r.get(find_col(r, ["book_title", "book", "书名"]) or "", "").strip()
    isbn = r.get(find_col(r, ["isbn"]) or "", "").strip()
    publisher = r.get(find_col(r, ["publisher", "出版社"]) or "", "").strip()
    year = r.get(find_col(r, ["year", "年份"]) or "", "").strip()
    
    # Build the citation like the example
    parts: List[str] = []
    if authors:
        parts.append(authors)
    if chapter_title:
        parts.append(f'Chapter "{chapter_title}"')
    if book_title:
        parts.append(f'for Book "{book_title}"')
    if isbn:
        parts.append(f"ISBN: {isbn}")
    if publisher:
        parts.append(f"<i>{publisher}</i>")
    if year:
        parts.append(year)
    
    main_text = ", ".join(parts)
    if main_text and not main_text.endswith("."):
        main_text += "."
    
    # Handle note and links
    note_key = find_col(r, ["note", "备注"])
    note_val = (r.get(note_key, "") if note_key else "").strip()
    
    note_link_key = find_col(r, ["note_link", "备注链接", "note_url"])
    note_link_val = (r.get(note_link_key, "") if note_link_key else "").strip()
    
    link_key = find_col(r, ["url", "link", "链接", "地址"]) or ""
    link_val = r.get(link_key, "").strip() if link_key else ""
    
    doi_key = find_col(r, ["doi", "doa"]) or ""
    doi_val = r.get(doi_key, "").strip() if doi_key else ""
    
    # Handle image from pics column
    image_key = find_col(r, ["pics", "image", "图片", "picture"]) or ""
    image_val = r.get(image_key, "").strip() if image_key else ""
    
    extras: List[str] = []
    if note_val:
        if note_link_val:
            # Make note a hyperlink if note_link exists
            extras.append(f"<font class=\"red\">(<a href='{note_link_val}' class=\"red\">{note_val}</a>)</font>")
        else:
            # Regular note without link
            extras.append(f"<font class=\"red\">({note_val})</font>")
    
    link_html = ""
    if link_val:
        link_html = f"<br><a href='{link_val}' class=\"deepgrey\">{link_val}</a>"
    elif doi_val:
        doi_url = f"https://doi.org/{doi_val}"
        link_html = f"<br><a href='{doi_url}' class=\"deepgrey\">{doi_url}</a>"
    
    image_html = ""
    if image_val:
        image_html = f'<br><br><img src="{image_val}" width="20%" class="animate__animated animate__fadeIn animate__slow animate__delay-1s">'
    
    extras_text = (" " + " ".join(extras)) if extras else ""
    return f"<li>{main_text}{extras_text}{link_html}{image_html}</li>"


def build_pub_items(rows: List[Dict[str, str]]) -> str:
    # Ensure a blank line after every item for better readability
    return "\n\n".join(build_pub_item(r) for r in rows)


def build_featured_items(rows: List[Dict[str, str]]) -> str:
    # Each featured item ends with an explicit <br>
    return "\n\n".join(build_pub_item_with_trailing_br(r) for r in rows)


def build_book_chapter_items(rows: List[Dict[str, str]]) -> str:
    # Each book chapter ends with explicit spacing
    return "\n\n".join(build_book_chapter_item(r) + "\n<br>" for r in rows)


def build_conference_items_with_trailing_br(rows: List[Dict[str, str]]) -> str:
    # Each conference item ends with <br>
    return "\n\n".join(build_pub_item(r) + "\n<br>" for r in rows)


def is_hospital_related(r: Dict[str, str]) -> bool:
    """Heuristic: check if any key/value indicates hospital/clinic context."""
    keywords = [
        "hospital", "clinic", "clinical", "medical", "medicine", "医学", "医院", "临床"
    ]
    hay = " ".join([f"{k} {v}" for k, v in r.items() if v]).lower()
    return any(kw in hay for kw in keywords)


def build_journal_items_with_trailing_br(rows: List[Dict[str, str]]) -> str:
    out: List[str] = []
    for r in rows:
        item = build_pub_item(r) + "\n<br>"
        out.append(item)
    return "\n\n".join(out)


def build_simple_list(rows: List[Dict[str, str]]) -> str:
    """Fallback: join non-empty row values per item for generic lists (Awards, Patents)."""
    items: List[str] = []
    for r in rows:
        text = ", ".join(v for v in r.values() if v)
        if not text:
            continue
        items.append(f"<li>{text}</li>")
        items.append("<br>")
    return "\n".join(items)


def strip_leading_index(text: str) -> str:
    # Remove leading numbering like "1, ", "1. ", "1) " etc.
    return re.sub(r"^\s*\d+\s*([,\.)、]\s*)?", "", text)


def build_patent_items(rows: List[Dict[str, str]]) -> str:
    items: List[str] = []
    for r in rows:
        # Compose from typical fields; ignore index/visible/note
        idx_key = find_col(r, ["index", "序号"]) or ""
        state_key = find_col(r, ["state", "状态"]) or ""
        visible_key = find_col(r, ["visible", "visable", "展示"]) or ""

        # Skip by visible flag if explicitly 0
        if visible_key:
            vis_raw = str(r.get(visible_key, "")).strip()
            if vis_raw and vis_raw not in {"1", "true", "True", "是", "yes", "Yes"}:
                continue

        # Build main text: title first, then other fields (excluding authors)
        title_key = find_col(r, ["title", "题目", "patent_title"]) or ""
        title_val = r.get(title_key, "").strip() if title_key else ""
        
        authors_key = find_col(r, ["authors", "author", "作者"]) or ""
        
        parts: List[str] = []
        if title_val:
            parts.append(strip_leading_index(title_val))
        
        # Add other fields except control/special ones and authors
        for k, v in r.items():
            if not v:
                continue
            kl = k.lower()
            if kl in {"", idx_key.lower(), state_key.lower(), visible_key.lower(), "note", "备注", "author_order", "作者序", "作者次序", "次序", title_key.lower() if title_key else "", authors_key.lower() if authors_key else ""}:
                continue
            parts.append(v)
        
        text = ", ".join(parts)

        # Append state in parentheses with color coding: 公开=blue, 授权=red
        state_val = r.get(state_key, "").strip() if state_key else ""
        state_html = ""
        if state_val:
            if "公开" in state_val:
                state_html = f" <font class=\"blue\"><b>({state_val})</b></font>"
            elif "授权" in state_val:
                state_html = f" <font class=\"red\"><b>({state_val})</b></font>"
            else:
                # Default blue for other states
                state_html = f" <font class=\"blue\"><b>({state_val})</b></font>"

        if text:
            items.append(f"<li>{text}{state_html}</li>")
            items.append("<br>")
    return "\n".join(items)


def build_award_items(rows: List[Dict[str, str]]) -> str:
    items: List[str] = []
    for r in rows:
        # Filter by visible/visable if present
        visible_key = find_col(r, ["visible", "visable", "展示"]) or ""
        if visible_key:
            vis_raw = str(r.get(visible_key, "")).strip()
            if vis_raw and vis_raw not in {"1", "true", "True", "是", "yes", "Yes"}:
                continue

        # Build content, excluding index and control fields
        idx_key = find_col(r, ["index", "序号"]) or ""
        note_key = find_col(r, ["note", "备注"]) or ""
        year_key = find_col(r, ["year", "年份"]) or ""
        time_key = find_col(r, ["time", "日期", "时间"]) or ""
        parts: List[str] = []
        for k, v in r.items():
            if not v:
                continue
            kl = k.lower()
            if kl in {"", idx_key.lower(), (visible_key.lower() if visible_key else ""), note_key.lower() if note_key else ""}:
                continue
            parts.append(v)
        text = ", ".join(parts)
        text = strip_leading_index(text)

        # Append note in red parentheses if present
        note_val = r.get(note_key, "").strip() if note_key else ""
        note_html = f" <font class=\"red\">({note_val})</font>" if note_val else ""

        # Prepend time if available, else year; format like "2024, ..."
        time_val = r.get(time_key, "").strip() if time_key else ""
        year_val = r.get(year_key, "").strip() if year_key else ""
        prefix_val = time_val or year_val
        if prefix_val:
            text = f"{prefix_val}, {text}"

        if text:
            items.append(f"<li>{text}{note_html}</li>")
            items.append("<br>")
    return "\n".join(items)


def build_correspondence(rows: List[Dict[str, str]]) -> str:
    # Handle two possible formats:
    # 1. Traditional: email/office/address/phone/homepage columns
    # 2. Name-Content: rows with "name" and "content" columns
    
    if not rows:
        return ""
    
    # Check if we have name/content structure
    first_row = rows[0]
    name_key = find_col(first_row, ["name", "名称", "field"])
    content_key = find_col(first_row, ["content", "内容", "value"])
    
    items: List[str] = []
    
    if name_key and content_key:
        # Name-Content format: each row is a field
        for r in rows:
            name_val = r.get(name_key, "").strip()
            content_val = r.get(content_key, "").strip()
            if name_val and content_val:
                items.append(f"<li><i>{name_val}</i> :&nbsp;&nbsp; {content_val}</li>")
                items.append("<br>")
    else:
        # Traditional format: use first non-empty row with field columns
        row = next((r for r in rows if any(v for v in r.values())), {})
        email = row.get(find_col(row, ["email", "邮箱", "e-mail"]) or "", "").strip()
        office = row.get(find_col(row, ["office", "办公室", "lab", "laboratory"]) or "", "").strip()
        address = row.get(find_col(row, ["address", "地址", "location"]) or "", "").strip()
        phone = row.get(find_col(row, ["phone", "电话", "tel"]) or "", "").strip()
        homepage = row.get(find_col(row, ["homepage", "website", "个人主页"]) or "", "").strip()

        if email:
            items.append(f"<li><i>Email</i> :&nbsp;&nbsp; {email}</li>")
            items.append("<br>")
        if phone:
            items.append(f"<li><i>Phone</i> :&nbsp;&nbsp; {phone}</li>")
            items.append("<br>")
        if office:
            items.append(f"<li><i>Office</i> :&nbsp;&nbsp; {office}</li>")
            items.append("<br>")
        if address:
            items.append(f"<li><i>Address</i> :&nbsp;&nbsp; {address}</li>")
            items.append("<br>")
        if homepage:
            items.append(f"<li><i>Homepage</i> :&nbsp;&nbsp; <a href='{homepage}' class=\"deepgrey\">{homepage}</a></li>")
    
    return "\n".join(items)


def build_text_block(rows: List[Dict[str, str]]) -> str:
    # Join all non-empty cells in first non-empty row
    for r in rows:
        vals = [v for v in r.values() if v]
        if vals:
            return " ".join(vals)
    return ""


def replace_section_list(html: str, section_id: str, list_tag: str, new_inner: str) -> str:
    """Replace the inner HTML of the first list (ul/ol) after a given section div id.

    Preserves the opening/closing tag and attributes, and attempts to keep indentation."""
    anchor = f'<div id="{section_id}"></div>'
    start_anchor = html.find(anchor)
    if start_anchor == -1:
        return html

    # Search for the first target list after the anchor
    open_tag_re = re.compile(rf"<\s*{list_tag}([^>]*)>", re.IGNORECASE)
    m_open = open_tag_re.search(html, pos=start_anchor)
    if not m_open:
        return html
    open_start = m_open.start()
    open_end = m_open.end()

    close_tag_re = re.compile(rf"</\s*{list_tag}\s*>", re.IGNORECASE)
    m_close = close_tag_re.search(html, pos=open_end)
    if not m_close:
        return html
    close_start = m_close.start()
    close_end = m_close.end()

    # Determine indentation based on current first item line
    current_inner = html[open_end:close_start]
    # Extract indentation from the first non-empty line inside
    indent_match = re.search(r"\n([ \t]+)\S", current_inner)
    indent = indent_match.group(1) if indent_match else "            "  # default spaces

    # Normalize new inner with indentation
    # Add initial <br> for spacing, then content
    inner_lines = [line for line in new_inner.splitlines()]
    if inner_lines:
        new_inner_indented = "\n" + indent + "<br>\n" + "\n".join(indent + line for line in inner_lines) + "\n"
    else:
        new_inner_indented = "\n" + indent + "<br>\n"

    return html[:open_end] + new_inner_indented + html[close_start:]


def replace_text_after_section(html: str, section_id: str, tag: str, new_text: str) -> str:
    """Replace the inner text of the first tag (e.g., h5) after a section anchor."""
    anchor = f'<div id="{section_id}"></div>'
    start_anchor = html.find(anchor)
    if start_anchor == -1:
        return html

    open_tag_re = re.compile(rf"<\s*{tag}([^>]*)>", re.IGNORECASE)
    m_open = open_tag_re.search(html, pos=start_anchor)
    if not m_open:
        return html
    open_end = m_open.end()

    close_tag_re = re.compile(rf"</\s*{tag}\s*>", re.IGNORECASE)
    m_close = close_tag_re.search(html, pos=open_end)
    if not m_close:
        return html

    # Preserve indentation around text content
    before = html[:open_end]
    after = html[m_close.start():]

    # Determine indentation from existing block
    existing_inner = html[open_end:m_close.start()]
    indent_match = re.search(r"\n([ \t]+)\S", existing_inner)
    indent = indent_match.group(1) if indent_match else "            "
    new_inner = f"\n{indent}{new_text}\n"
    return before + new_inner + after


def process_workbook_to_html(xlsx_path: Path, html_path: Path) -> None:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    html = html_path.read_text(encoding="utf-8")

    # Build updates per section
    updates: List[Tuple[str, str, str]] = []  # (section_id, tag, new_inner)
    text_updates: List[Tuple[str, str, str]] = []  # (section_id, tag, new_text)

    for ws in wb.worksheets:
        sheet_name_norm = normalize_name(ws.title)
        rows = read_rows_as_dicts(ws)
        if not rows:
            continue

        # Skip Biography explicitly
        if sheet_name_norm in {"biography", "bio", "传记", "简历"}:
            continue

        # News
        if sheet_name_norm in {"news", "最新动态", "动态"}:
            updates.append(("News", "ul", build_news_items(rows)))
            continue

        # Publications
        if sheet_name_norm in {"featured article", "featured", "精选文章"}:
            updates.append(("Featured Article", "ol", build_featured_items(rows)))
            continue
        if sheet_name_norm in {"book chapter", "book chapters", "图书章节"}:
            updates.append(("Book Chapter", "ol", build_book_chapter_items(rows)))
            continue
        if sheet_name_norm in {"refereed journal papers", "journal", "journals", "期刊论文"}:
            updates.append(("Refereed Journal Papers", "ol", build_journal_items_with_trailing_br(rows)))
            continue
        if sheet_name_norm in {"refereed conference papers", "refereed conference paper", "conference", "conferences", "会议论文", "conf", "conference papers", "refereed conferences", "会议"}:
            updates.append(("Refereed Conference Papers", "ol", build_conference_items_with_trailing_br(rows)))
            continue

        # Patents
        if sheet_name_norm in {"patent", "patents", "专利"}:
            updates.append(("Patent", "ol", build_patent_items(rows)))
            continue

        # Awards
        if sheet_name_norm in {"award", "awards", "award & honors", "honors", "荣誉", "奖项"}:
            updates.append(("Award", "ul", build_award_items(rows)))
            continue

        # Correspondence
        if sheet_name_norm in {"correspondence", "contact", "contacts", "联系方式", "联系", "通讯"}:
            updates.append(("Correspondence", "ul", build_correspondence(rows)))
            continue

        # Research Interests
        if sheet_name_norm in {"research interests", "interests", "研究兴趣"}:
            text_updates.append(("Research Interests", "h5", build_text_block(rows)))
            continue

        # Header/Profile (optional). Update the top name and profile block if provided.
        if sheet_name_norm in {"header", "profile", "顶部", "基本信息"}:
            # Attempt to update name, english name, titles, email, lab, ieee link
            row = rows[0]
            name_cn = row.get(find_col(row, ["name_cn", "姓名"]) or "", "")
            name_en = row.get(find_col(row, ["name_en", "英文名"]) or "", "")
            title_lines = [
                row.get(find_col(row, ["title1"]) or "", "").strip(),
                row.get(find_col(row, ["title2"]) or "", "").strip(),
                row.get(find_col(row, ["title3"]) or "", "").strip(),
            ]
            email = row.get(find_col(row, ["email"]) or "", "")
            ieee = row.get(find_col(row, ["ieee", "ieee link"]) or "", "")
            # Replace simple occurrences in the header block by regex if present
            if name_cn:
                html = re.sub(r"(<h2 class=\"white\">)(.*?)(</h2>)", rf"\1{name_cn}\3", html, count=1)
            if name_en:
                html = re.sub(r"(<h3 class=\"white\">)(.*?)(</h3>)", rf"\1{name_en}\3", html, count=1)
            if email:
                html = re.sub(r"(<h5 class=\"white\">E-mail:)(.*?)(</h5>)", rf"\1 {email}\3", html, count=1)
            if ieee:
                html = re.sub(
                    r"(<a href=\").*?(\" class=\"blue\">IEEE author homepage</a>)",
                    rf"\1{ieee}\2",
                    html,
                    count=1,
                )
            continue

        # Unknown sheets are ignored gracefully

    # Apply list updates
    for section_id, tag, new_inner in updates:
        html = replace_section_list(html, section_id=section_id, list_tag=tag, new_inner=new_inner)

    # Apply text updates
    for section_id, tag, new_text in text_updates:
        html = replace_text_after_section(html, section_id=section_id, tag=tag, new_text=new_text)

    html_path.write_text(html, encoding="utf-8")


def main():
    start_time = time.time()
    
    xlsx_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_XLSX_PATH
    html_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_HTML_PATH

    if not xlsx_path.exists():
        print(f"Excel file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    if not html_path.exists():
        print(f"HTML file not found: {html_path}", file=sys.stderr)
        sys.exit(1)

    process_workbook_to_html(xlsx_path, html_path)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Updated {html_path} from {xlsx_path}")
    print(f"Execution time: {elapsed_time:.2f} seconds")


if __name__ == "__main__":
    main()


